import os
import time
import sys
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import logging
import subprocess

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

if len(sys.argv) != 3:
    raise Exception("Erro: Login e senha de administrador não fornecidos.")
admin_username = sys.argv[1]
admin_password = sys.argv[2]

# Função para criar o diretório de downloads
def criar_diretorio_download():
    diretorio_download = os.path.join(os.getcwd(), "downloads")
    if not os.path.exists(diretorio_download):
        os.makedirs(diretorio_download)
    return diretorio_download

# Função para configurar o navegador
def configurar_navegador(diretorio_download, headless=False):
    chrome_options = webdriver.ChromeOptions()
    prefs = {"download.default_directory": diretorio_download}
    chrome_options.add_experimental_option("prefs", prefs)
    
    if headless:
        chrome_options.add_argument("--headless")  # Executa o navegador em modo headless
        chrome_options.add_argument("--disable-gpu")  # Necessário para o modo headless funcionar corretamente em alguns sistemas
        chrome_options.add_argument("--window-size=1920,1080")  # Define um tamanho de janela padrão

    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    return navegador


# Função para fazer login
def fazer_login(navegador, username, password):
    max_retries = 3
    for attempt in range(max_retries):
        try:
            navegador.get("https://santacasabahia.mentorlearn.com/login")
            navegador.maximize_window()
            campo_user = WebDriverWait(navegador, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="username"]'))
            )
            campo_user.send_keys(username)
            campo_senha = WebDriverWait(navegador, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="password"]'))
            )
            campo_senha.send_keys(password)
            clicar_com_javascript(navegador, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-login/div/div/div[2]/form/div/div[3]/p/button')
            WebDriverWait(navegador, 20).until(EC.url_changes("https://santacasabahia.mentorlearn.com/login"))
            return True
        except Exception as e:
            logging.error(f"Erro ao tentar fazer login com {username} na tentativa {attempt + 1}")
            time.sleep(5)  # Esperar alguns segundos antes de tentar novamente
    return False

# Função para clicar usando JavaScript
def clicar_com_javascript(navegador, xpath, tempo_espera=0):
    try:
        elemento = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        navegador.execute_script("arguments[0].click();", elemento)
        if tempo_espera > 0:
            time.sleep(tempo_espera)
    except Exception as e:
        logging.error(f"Erro ao clicar no elemento {xpath}")

# Função para extrair informações necessárias e adicionar ao Excel
def extrair_informacoes(navegador, sheet):
    try:
        campo_input = navegador.find_element(By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav[2]/div/ng-component/ml-users-groups-settings/mat-tab-group/div/mat-tab-body[3]/div/ml-users-groups-settings-details/div/mat-form-field[3]/div/div[1]/div/input')
        loginUser = campo_input.get_attribute('value')
    except Exception as e:
        logging.error(f"Erro ao extrair informações:")
        loginUser = "N/A"
    sheet.append([loginUser])

# Função para gerar o primeiro relatório
def gerar_relatorio_1(navegador, username, diretorio_download):
    try:
        navegador.get('https://santacasabahia.mentorlearn.com/reports/sim_typ8(secondary:reports-settings)')
        campo_data = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav[2]/div/ng-component/ml-report-settings/div/mat-accordion/mat-expansion-panel[1]/div/div/div/div/mat-form-field[1]/div/div[1]/div[1]/input'))
        )
        campo_data.send_keys(Keys.CONTROL + "a")
        campo_data.send_keys(Keys.DELETE)
        campo_data.send_keys('1/1/2021')
        campo_data.send_keys(Keys.TAB)
        
        time.sleep(5)  # Espera fazer a busca de todos os itens

        if WebDriverWait(navegador, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-reports/div/mat-table/mat-row[1]'))):
            clicar_com_javascript(navegador, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-reports/div/mat-table/mat-header-row/mat-header-cell[1]/mat-checkbox/label', 2)
            clicar_com_javascript(navegador, '/html/body/ml-app/div/ml-toolbar/div/mat-toolbar/div[5]/ml-toolbar-more-menu/button', 2)
            clicar_com_javascript(navegador, '/html/body/div/div[2]/div/div/div/button', 2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/mat-checkbox/label', 2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/div[2]/mat-radio-group/mat-radio-button[2]/label',2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[3]/div[2]/button', 2)

            # Verificar até que o ícone de carregamento desapareça
            WebDriverWait(navegador, 240).until(
                EC.invisibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/div[3]'))
            )
            time.sleep(10)  # Espera adicional após o desaparecimento do ícone

            # Verificação contínua da existência do arquivo baixado e renomeado
            tempo_limite = time.time() + 60  # 60 segundos de espera
            novo_nome = os.path.join(diretorio_download, f"{username}_report8.zip")
            while time.time() < tempo_limite:
                try:
                    ultimo_arquivo = max([os.path.join(diretorio_download, f) for f in os.listdir(diretorio_download)], key=os.path.getctime)
                    os.rename(ultimo_arquivo, novo_nome)
                    break
                except (ValueError, FileNotFoundError):
                    time.sleep(1)

            if not os.path.exists(novo_nome):
                logging.error(f"Fazendo a troca de usuario")
        else:
            logging.warning(f"O elemento para selecionar todos os itens do relatório não foi encontrado para {username}. Não pode ser gerado um relatorio")

    except Exception as e:
        logging.error(f"Erro ao gerar o relatório 8 para {username}: {e}")

# Função para gerar o segundo relatório
def gerar_relatorio_2(navegador, username, diretorio_download):
    try:
        navegador.get('https://santacasabahia.mentorlearn.com/reports/sim_typ1(secondary:reports-settings)')
        campo_data = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav[2]/div/ng-component/ml-report-settings/div/mat-accordion/mat-expansion-panel[1]/div/div/div/div/mat-form-field[1]/div/div[1]/div[1]/input'))
        )
        campo_data.send_keys(Keys.CONTROL + "a")
        campo_data.send_keys(Keys.DELETE)
        campo_data.send_keys('1/1/2021')
        campo_data.send_keys(Keys.TAB)
        
        time.sleep(5)  # Espera fazer a busca de todos os itens

        if WebDriverWait(navegador, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-reports/div/mat-table/mat-row[1]'))):
            clicar_com_javascript(navegador, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-reports/div/mat-table/mat-header-row/mat-header-cell[1]/mat-checkbox/label', 2)
            clicar_com_javascript(navegador, '/html/body/ml-app/div/ml-toolbar/div/mat-toolbar/div[5]/ml-toolbar-more-menu/button', 2)
            clicar_com_javascript(navegador, '/html/body/div/div[2]/div/div/div/button', 2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/mat-checkbox/label', 2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/div[2]/mat-radio-group/mat-radio-button[2]/label',2)
            clicar_com_javascript(navegador, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[3]/div[2]/button', 2)

            # Verificar até que o ícone de carregamento desapareça
            WebDriverWait(navegador, 240).until(
                EC.invisibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/mat-dialog-container/ml-dynamic-dialog/div/div[2]/ml-export-performance/div/div[3]'))
            )
            time.sleep(10)  # Espera adicional após o desaparecimento do ícone

            # Verificação contínua da existência do arquivo baixado e renomeado
            tempo_limite = time.time() + 60  # 60 segundos de espera
            novo_nome = os.path.join(diretorio_download, f"{username}_report1.zip")
            while time.time() < tempo_limite:
                try:
                    ultimo_arquivo = max([os.path.join(diretorio_download, f) for f in os.listdir(diretorio_download)], key=os.path.getctime)
                    os.rename(ultimo_arquivo, novo_nome)
                    break
                except (ValueError, FileNotFoundError):
                    time.sleep(1)

            if not os.path.exists(novo_nome):
                logging.error(f"Arquivo não encontrado para o relatório 1 de {username} após o tempo limite.")
                navegador.close()  # Fecha a aba se o relatório não for encontrado
                navegador.switch_to.window(navegador.window_handles[0])  # Voltar para a aba original
        else:
            logging.warning(f"O elemento para selecionar todos os itens do relatório 1 não foi encontrado para {username}")

    except Exception as e:
        logging.error(f"Erro ao gerar o relatório 1 para {username}")

# Função principal
def main():
    diretorio_download = criar_diretorio_download()

    # Navegador inicial para login do administrador
    navegador = configurar_navegador(diretorio_download)

    try:
        if not fazer_login(navegador, admin_username, admin_password):
            raise Exception("Erro ao fazer login com a conta de administrador")

        navegador.get('https://santacasabahia.mentorlearn.com/users-groups(secondary:users-groups-settings)')
        expandir_pronon = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-users-groups/div/div/ml-users-groups-view/div/mat-tree/mat-tree-node[4]/div/button'))
        )
        expandir_pronon.click()

        if os.path.exists("usuarios.xlsx"):
            pass
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "usuarios"
            sheet.append(["loginUser"])

            for i in range(12, 220):
                expandir_xpath = f'/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav-content/ml-users-groups/div/div/ml-users-groups-view/div/mat-tree/mat-tree-node[{i}]'
                try:
                    clicar_com_javascript(navegador, expandir_xpath)
                    time.sleep(2)

                    profile_xpath = f'/html/body/ml-app/div/ml-main-content/mat-sidenav-container/mat-sidenav[2]/div/ng-component/ml-users-groups-settings/mat-tab-group/mat-tab-header/div[2]/div/div/div[3]'
                    profile = WebDriverWait(navegador, 20).until(
                        EC.presence_of_element_located((By.XPATH, profile_xpath))
                    )
                    profile.click()
                    extrair_informacoes(navegador, sheet)
                    time.sleep(1)

                    clicar_com_javascript(navegador, expandir_xpath)
                    time.sleep(1)

                except Exception as e:
                    logging.error(f"Erro ao clicar ou extrair informações do nó {i}")

            workbook.save("usuarios.xlsx")

        # Fecha o navegador após obter as informações dos usuários
        navegador.quit()

        # Reabre o navegador para cada usuário
        workbook = openpyxl.load_workbook("usuarios.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username = row[0]
            password = username

            navegador = configurar_navegador(diretorio_download)  # Reabre o navegador
            if fazer_login(navegador, username, password):
                gerar_relatorio_1(navegador, username, diretorio_download)
                navegador.quit()  # Fecha o navegador após gerar os relatórios
            else:
                navegador.quit()
                continue

        workbook = openpyxl.load_workbook("usuarios.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username = row[0]
            password = username

            navegador = configurar_navegador(diretorio_download)  # Reabre o navegador
            if fazer_login(navegador, username, password):
                gerar_relatorio_2(navegador, username, diretorio_download)
                navegador.quit()  # Fecha o navegador após gerar os relatórios
            else:
                navegador.quit()
                continue
    
    # Inicia o TratamentoZIP.py
        logging.info("Iniciando o TratamentoZIP.py...")
        resultado_zip = subprocess.run(["python", "TratamentoZIPparcial.py"])

        if resultado_zip.returncode == 0:
            logging.info("TratamentoZIP.py concluído com sucesso.")
            
            # Agora inicia o TratamentoCSV.py
            logging.info("Iniciando o TratamentoCSVparcial.py...")
            resultado_csv = subprocess.run(["python", "TratamentoCSVparcial.py"])

            if resultado_csv.returncode == 0:
                logging.info("TratamentoCSV.py concluído com sucesso.")
            else:
                logging.error("TratamentoCSVparcial.py falhou.")
        else:
            logging.error("TratamentoZIP.py falhou. TratamentoCSVparcial.py não será iniciado.")


    finally:
        navegador.quit()

if __name__ == "__main__":
    main()